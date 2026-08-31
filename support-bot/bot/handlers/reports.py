"""
Экспорт и отчёты: /export (CSV), /report, /report_pdf, /report_excel, /dashboard.

Исправление: CSV/Excel-ячейки экранируются через csv_safe() — защита от
"CSV injection" (пользовательский текст, начинающийся с =, +, -, @,
может быть распознан Excel/Google Sheets как формула при открытии файла).
"""
import io
import logging
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from telegram import InputFile, Update
from telegram.ext import ContextTypes

from bot.db import repository as repo
from bot.utils import csv_safe

logger = logging.getLogger(__name__)


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    import csv

    days = int(context.args[0]) if context.args and context.args[0].isdigit() else None
    rows = await repo.generate_export(days)
    if not rows:
        await update.message.reply_text("Нет данных.")
        return
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Автор", "Юзернейм", "Текст", "Тип", "Статус", "Приоритет", "Теги", "Ответственные", "Создано"])
    for row in rows:
        writer.writerow([csv_safe(v) for v in row])
    output.seek(0)
    await update.message.reply_document(
        document=output.getvalue().encode("utf-8"),
        filename=f"tasks_export_{datetime.now().strftime('%Y%m%d')}.csv",
        caption="📊 Выгрузка",
    )


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_created, total_closed, type_counts, priority_counts = await repo.generate_weekly_report()
    response = f"📈 Отчёт за 7 дней:\n📌 Создано: {total_created}\n✅ Закрыто: {total_closed}\n\nПо типам:\n"
    for t, cnt in type_counts:
        response += f"  {t}: {cnt}\n"
    response += "\nПо приоритетам:\n"
    priority_names = {"high": "🔴 Критичный", "medium": "🟡 Важный", "low": "🟢 Обычный"}
    for p, cnt in priority_counts:
        response += f"  {priority_names.get(p, p)}: {cnt}\n"
    await update.message.reply_text(response)


async def report_pdf_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_created, total_closed, type_counts, priority_counts = await repo.generate_weekly_report()
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.drawString(100, height - 50, "Еженедельный отчёт по задачам")
    c.drawString(100, height - 80, f"Создано за неделю: {total_created}")
    c.drawString(100, height - 100, f"Закрыто за неделю: {total_closed}")
    y = height - 130
    c.drawString(100, y, "По типам:")
    y -= 20
    for t, cnt in type_counts:
        c.drawString(120, y, f"{t}: {cnt}")
        y -= 20
    y -= 10
    c.drawString(100, y, "По приоритетам:")
    y -= 20
    for p, cnt in priority_counts:
        c.drawString(120, y, f"{p}: {cnt}")
        y -= 20
    c.save()
    buffer.seek(0)
    await update.message.reply_document(
        document=InputFile(buffer, filename="weekly_report.pdf"),
        caption="📄 Еженедельный отчёт в PDF",
    )


def _build_excel_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Решённые задачи"
    headers = ["ID", "Заголовок", "Текст", "Автор", "Ответственный", "Создано", "Закрыто", "Приоритет", "Теги"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            row[0], csv_safe(row[1]), csv_safe(row[2]), csv_safe(row[3]), csv_safe(row[4]),
            row[5] or "", row[6] or "", csv_safe(row[7]), csv_safe(row[8]),
        ])
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_length = max((len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None), default=0)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    ws2 = wb.create_sheet("Сводка по ответственным")
    ws2.append(["Ответственный", "Количество закрытых задач"])
    from collections import Counter
    counter = Counter(row[4] if row[4] else "Не назначен" for row in rows)
    for resp, count in counter.most_common():
        ws2.append([resp, count])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def report_excel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /report_excel [from YYYY-MM-DD] [to YYYY-MM-DD]"""
    args = context.args
    from_date = to_date = None
    for i, arg in enumerate(args):
        if arg.lower() == "from" and i + 1 < len(args):
            try:
                from_date = datetime.strptime(args[i + 1], "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        elif arg.lower() == "to" and i + 1 < len(args):
            try:
                to_date = datetime.strptime(args[i + 1], "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
    if not from_date and not to_date:
        from_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
        to_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    await update.message.reply_text("📊 Генерирую Excel-отчёт...")
    try:
        rows = await repo.generate_excel_rows(from_date, to_date)
        excel_file = _build_excel_workbook(rows)
        filename = f"report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        await update.message.reply_document(
            document=excel_file,
            filename=filename,
            caption=f"📋 Отчёт по закрытым задачам с {from_date[:10] if from_date else 'начала'} по {to_date[:10] if to_date else 'конец'}",
        )
    except Exception as e:
        logger.error(f"Ошибка генерации Excel-отчёта: {e}")
        await update.message.reply_text("❌ Не удалось сгенерировать отчёт. Проверьте логи.")


async def dashboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_created, total_closed, priority_counts, top_tags, top_users = await repo.get_dashboard_data()

    html = [
        "<html><head><meta charset='utf-8'><title>Дашборд</title></head><body>",
        "<h1>📊 Дашборд задач</h1>",
        f"<p><b>Создано за неделю:</b> {total_created}</p>",
        f"<p><b>Закрыто за неделю:</b> {total_closed}</p>",
        "<h2>Приоритеты</h2><ul>",
    ]
    for p, cnt in priority_counts:
        html.append(f"<li>{p}: {cnt}</li>")
    html.append("</ul><h2>Топ-5 тегов</h2><ul>")
    for tag, cnt in top_tags:
        html.append(f"<li>{tag}: {cnt}</li>")
    html.append("</ul><h2>Топ-5 пользователей</h2><ul>")
    for uid, pts in top_users:
        html.append(f"<li>ID {uid}: {pts} очков</li>")
    html.append("</ul></body></html>")

    html_bytes = "".join(html).encode("utf-8")
    await update.message.reply_document(
        document=io.BytesIO(html_bytes),
        filename="dashboard.html",
        caption="📈 Дашборд (откройте в браузере)",
    )
